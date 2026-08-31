import datetime
import json
import math

from openpyxl import load_workbook

MAX_SHEETS = 200
MAX_ROWS = 100_000
MAX_COLUMNS = 16_384
MAX_VISITED_CELLS = 200_000
MAX_SOURCES = 100_000
MAX_SOURCE_CODE_POINTS = 100_000
MAX_DOCUMENT_CODE_POINTS = 2_000_000


def value_text(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float):
        if not math.isfinite(value):
            raise ValueError("number_invalid")
        return repr(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    raise ValueError("cell_type_unsupported")


def add_source(sources, sheet_index, cell, content, state):
    if len(content) > MAX_SOURCE_CODE_POINTS:
        raise ValueError("source_limit")
    state["points"] += len(content)
    if state["points"] > MAX_DOCUMENT_CODE_POINTS:
        raise ValueError("document_limit")
    if len(sources) >= MAX_SOURCES:
        raise ValueError("source_count_limit")
    sources.append({"sheet": sheet_index, "cell": cell.coordinate, "content": content})


def add_sheet(worksheet, sheet_index, sources, state):
    max_row = worksheet.max_row
    max_column = worksheet.max_column
    if max_row > MAX_ROWS or max_column > MAX_COLUMNS:
        raise ValueError("sheet_dimension_limit")
    if max_row * max_column > MAX_VISITED_CELLS:
        raise ValueError("cell_visit_limit")
    for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_column):
        for cell in row:
            state["visited"] += 1
            if state["visited"] > MAX_VISITED_CELLS:
                raise ValueError("cell_visit_limit")
            content = value_text(cell.value)
            if content:
                add_source(sources, sheet_index, cell, content, state)


def parse_xlsx(path):
    workbook = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    sources: list[dict[str, object]] = []
    state = {"visited": 0, "points": 0}
    try:
        for sheet_index, worksheet in enumerate(workbook.worksheets, start=1):
            if sheet_index > MAX_SHEETS:
                raise ValueError("sheet_count_limit")
            add_sheet(worksheet, sheet_index, sources, state)
    finally:
        workbook.close()
    if not sources:
        raise ValueError("no_text")
    return json.dumps(
        {"schema_version": "1", "format": "xlsx", "sources": sources},
        ensure_ascii=False,
        separators=(",", ":"),
    )


parse_xlsx("/tmp/aethelgard-source.xlsx")
